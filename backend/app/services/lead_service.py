import csv
import io
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models import Lead, LeadStatus, Call
from app.utils.phone import normalize_pk_phone


def _clean_value(v) -> str:
    if isinstance(v, list):
        v = " ".join(x for x in v if x)
    if v is None:
        return ""
    # openpyxl hands back phone-number-like columns as floats (e.g. 923001234567.0)
    # — stringify whole-number floats as plain integers or the trailing ".0" would
    # get merged into the digits by normalize_pk_phone's non-digit stripping.
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _header_row_index(rows: list[list[str]]) -> int:
    # Skip leading blank/junk rows (e.g. a stray ",," before the real header) so the
    # first row with at least 2 named columns is used as the header, not row 1 blindly.
    return next(
        (i for i, r in enumerate(rows) if sum(1 for cell in r if cell.strip()) >= 2),
        0,
    )


def _rows_from_csv(file_bytes: bytes):
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    lines = text.splitlines()
    all_rows = list(csv.reader(lines))
    header_idx = _header_row_index(all_rows)
    reader = csv.DictReader(lines[header_idx:])
    for row in reader:
        yield {(k or "").strip().lower(): _clean_value(v) for k, v in row.items()}


def _rows_from_excel(file_bytes: bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    all_rows = [
        [_clean_value(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    header_idx = _header_row_index(all_rows)
    if header_idx >= len(all_rows):
        return
    headers = [cell.strip().lower() for cell in all_rows[header_idx]]
    for row in all_rows[header_idx + 1:]:
        if not any(cell.strip() for cell in row):
            continue
        yield {
            headers[i] if i < len(headers) and headers[i] else f"col_{i}": row[i]
            for i in range(len(row))
        }


def import_leads_csv(db: Session, file_bytes: bytes, filename: str = "") -> dict:
    is_excel = filename.lower().endswith((".xlsx", ".xls"))
    rows = _rows_from_excel(file_bytes) if is_excel else _rows_from_csv(file_bytes)

    imported = 0
    duplicates = 0
    invalid = 0
    seen = set()
    existing = {p for (p,) in db.query(Lead.phone).all()}
    for row in rows:
        phone_key = next(
            (k for k in row if "phone" in k or "mobile" in k or "contact" in k or "number" in k),
            None,
        )
        phone = normalize_pk_phone(row.get(phone_key, "")) if phone_key else None
        if not phone:
            for value in row.values():
                phone = normalize_pk_phone(value)
                if phone:
                    break
        if not phone:
            invalid += 1
            continue
        if phone in seen or phone in existing:
            duplicates += 1
            continue
        seen.add(phone)
        lead = Lead(
            phone=phone,
            full_name=row.get("name") or row.get("full_name") or None,
            email=row.get("email") or None,
            program_of_interest=row.get("program") or row.get("program_of_interest") or None,
            crm_id=row.get("crm_id") or row.get("id") or None,
        )
        db.add(lead)
        imported += 1
    db.commit()
    return {"imported": imported, "duplicates_skipped": duplicates, "invalid_numbers": invalid}


TEST_LEAD_PHONE = "+923000000000"


def get_or_reset_test_lead(db: Session) -> Lead:
    """Single reusable lead for the browser Test Call page, so a real webhook round-trip
    can be verified end-to-end without needing a real phone call. Reused (not recreated)
    across test calls so it's always at /leads/<id> for the same id, and reset to a clean
    pending state each time so stale data from a previous test doesn't linger."""
    lead = db.query(Lead).filter(Lead.phone == TEST_LEAD_PHONE).first()
    if not lead:
        lead = Lead(phone=TEST_LEAD_PHONE)
        db.add(lead)
    lead.full_name = "Malaika"
    lead.status = LeadStatus.pending
    lead.retry_count = 0
    lead.next_retry_at = None
    lead.dnc = False
    lead.review_reason = None
    lead.reschedule_time = None
    lead.current_status = None
    lead.timeline = None
    lead.original_blocker = None
    lead.last_qualification = None
    lead.grade_or_cgpa = None
    lead.meets_baseline = None
    lead.advisor_callback_time = None
    lead.route_team = None
    db.commit()
    db.refresh(lead)
    return lead


def reset_all_leads(db: Session) -> int:
    # No dnc filter — resetting is now an explicit admin action that also clears a
    # mistaken DNC/hostile lock (e.g. a misclassified call), making every lead
    # callable again, not just non-DNC ones.
    updated = (
        db.query(Lead)
        .update(
            {
                Lead.status: LeadStatus.pending,
                Lead.retry_count: 0,
                Lead.next_retry_at: None,
                Lead.dnc: False,
                Lead.email: None,
                Lead.program_of_interest: None,
                Lead.wants_callback: False,
                Lead.reschedule_time: None,
                Lead.current_status: None,
                Lead.timeline: None,
                Lead.original_blocker: None,
                Lead.last_qualification: None,
                Lead.grade_or_cgpa: None,
                Lead.meets_baseline: None,
                Lead.advisor_callback_time: None,
                Lead.route_team: None,
                Lead.review_reason: None,
                Lead.field_confidence: None,
                Lead.reviewed_at: None,
                Lead.reviewed_by: None,
                Lead.crm_synced: False,
            },
            synchronize_session=False,
        )
    )
    db.commit()
    return updated


def get_stats(db: Session) -> dict:
    counts = dict(
        db.query(Lead.status, func.count(Lead.id)).group_by(Lead.status).all()
    )
    total_calls = db.query(func.count(Call.id)).scalar() or 0
    program_counts = (
        db.query(Lead.program_of_interest, func.count(Lead.id))
        .filter(Lead.program_of_interest.isnot(None), Lead.program_of_interest != "")
        .group_by(Lead.program_of_interest)
        .order_by(func.count(Lead.id).desc())
        .all()
    )
    return {
        "total": sum(counts.values()),
        "pending": counts.get(LeadStatus.pending, 0),
        "calling": counts.get(LeadStatus.calling, 0),
        "no_answer": counts.get(LeadStatus.no_answer, 0),
        "reactivated": counts.get(LeadStatus.reactivated, 0),
        "nurture": counts.get(LeadStatus.nurture, 0),
        "closed_lost": counts.get(LeadStatus.closed_lost, 0),
        "invalid": counts.get(LeadStatus.invalid, 0),
        "failed": counts.get(LeadStatus.failed, 0),
        "needs_review": counts.get(LeadStatus.needs_review, 0),
        "total_calls_made": total_calls,
        "by_program": [{"program": program, "count": count} for program, count in program_counts],
    }
